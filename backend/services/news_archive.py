import os
import glob
from typing import List, Dict, Optional
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

try:
    from models import NewsItem
except ImportError:
    pass

class BaseNewsArchive:
    """Abstract base class for news archiving system. Allows swapping storage backends later."""
    
    def archive_news(self, news_items: List['NewsItem']) -> int:
        raise NotImplementedError
        
    def get_by_date(self, date_str: str) -> List[Dict]:
        """date_str in YYYY-MM-DD format"""
        raise NotImplementedError

    def get_by_date_range(self, start_date: str, end_date: str) -> List[Dict]:
        """start_date, end_date in YYYY-MM-DD format"""
        raise NotImplementedError
        
    def get_available_dates(self) -> List[str]:
        raise NotImplementedError

class ExcelNewsArchive(BaseNewsArchive):
    def __init__(self, storage_dir: str):
        self.storage_dir = storage_dir
        os.makedirs(self.storage_dir, exist_ok=True)
        
        self.COLUMNS = [
            "timestamp", "date", "time", "title", "source", 
            "category", "sentiment", "related_symbols", "sentiment_score", "url"
        ]

    def _get_file_path(self, date_str: str) -> str:
        return os.path.join(self.storage_dir, f"news_{date_str}.xlsx")
        
    def _create_workbook_if_needed(self, filepath: str):
        if not os.path.exists(filepath):
            wb = Workbook()
            ws = wb.active
            ws.title = "News Archive"
            for col_idx, label in enumerate(self.COLUMNS, start=1):
                cell = ws.cell(row=1, column=col_idx, value=label)
                cell.font = Font(bold=True)
            wb.save(filepath)

    def archive_news(self, news_items: List['NewsItem']) -> int:
        if not news_items:
            return 0
            
        grouped = {}
        for item in news_items:
            if not item.published:
                continue
            date_str = item.published.strftime("%Y-%m-%d")
            if date_str not in grouped:
                grouped[date_str] = []
            grouped[date_str].append(item)
            
        total_saved = 0
        
        for date_str, items in grouped.items():
            filepath = self._get_file_path(date_str)
            self._create_workbook_if_needed(filepath)
            
            wb = load_workbook(filepath)
            ws = wb.active
            
            existing_hashes = set()
            for row in ws.iter_rows(min_row=2, values_only=True):
                title = str(row[3] or "").strip().lower()
                source = str(row[4] or "").strip().lower()
                existing_hashes.add(f"{title}|{source}")
                
            added = 0
            for item in items:
                title_lower = item.title.strip().lower()
                source_lower = item.source.strip().lower()
                item_hash = f"{title_lower}|{source_lower}"
                
                if item_hash in existing_hashes:
                    continue
                
                # Check for sentiment enum and extract string
                sentiment_val = getattr(item, "sentiment", "NEUTRAL")
                if hasattr(sentiment_val, "value"):
                    sentiment_val = sentiment_val.value
                sentiment_val = str(sentiment_val).upper().replace("SENTIMENTTYPE.", "")
                    
                row_data = [
                    item.published.strftime("%Y-%m-%d %H:%M:%S"),
                    item.published.strftime("%Y-%m-%d"),
                    item.published.strftime("%H:%M:%S"),
                    item.title.strip(),
                    item.source,
                    getattr(item, "category", "General"),
                    sentiment_val,
                    ",".join(getattr(item, "related_symbols", [])),
                    round(getattr(item, "sentiment_score", 0.0), 3),
                    item.url
                ]
                
                ws.append(row_data)
                existing_hashes.add(item_hash)
                added += 1
                
            if added > 0:
                wb.save(filepath)
                total_saved += added
                
        return total_saved

    def get_by_date(self, date_str: str) -> List[Dict]:
        filepath = self._get_file_path(date_str)
        if not os.path.exists(filepath):
            return []
            
        try:
            wb = load_workbook(filepath, read_only=True)
            ws = wb.active
            
            # Read header row to map column names to indexes
            header = []
            for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
                header = list(row) if row else []
                break
                
            col_map = {name: idx for idx, name in enumerate(header) if name is not None}
            
            output = []
            # Read data rows
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not any(row is not None for row in row):
                    continue
                
                def get_val(col_name, default=""):
                    idx = col_map.get(col_name)
                    if idx is not None and idx < len(row):
                        val = row[idx]
                        return val if val is not None else default
                    return default
                
                sentiment_val = str(get_val("sentiment", "NEUTRAL")).replace("SentimentType.", "")
                related = get_val("related_symbols", "")
                
                output.append({
                    "title": str(get_val("title", "")),
                    "source": str(get_val("source", "")),
                    "url": str(get_val("url", "")),
                    "published": str(get_val("timestamp", "")),
                    "sentiment": sentiment_val,
                    "sentiment_score": float(get_val("sentiment_score", 0.0)) if get_val("sentiment_score", 0.0) is not None else 0.0,
                    "category": str(get_val("category", "")),
                    "related_symbols": str(related).split(",") if related else []
                })
            
            output.sort(key=lambda x: x["published"], reverse=True)
            return output
        except Exception as e:
            import logging
            logging.error(f"Error reading archive {filepath}: {e}")
            return []

    def get_by_date_range(self, start_date: str, end_date: str) -> List[Dict]:
        try:
            start = datetime.strptime(start_date, "%Y-%m-%d")
            end = datetime.strptime(end_date, "%Y-%m-%d")
            delta = end - start
            dates = [(start + timedelta(days=i)).strftime("%Y-%m-%d") for i in range(delta.days + 1)]
        except Exception as date_err:
            import logging
            logging.error(f"Error parsing date range {start_date} to {end_date}: {date_err}")
            return []
            
        combined = []
        for d in dates:
            combined.extend(self.get_by_date(d))
        combined.sort(key=lambda x: x["published"], reverse=True)
        return combined

    def get_available_dates(self) -> List[str]:
        pattern = os.path.join(self.storage_dir, "news_*.xlsx")
        files = glob.glob(pattern)
        dates = []
        for f in files:
            basename = os.path.basename(f)
            d = basename.replace("news_", "").replace(".xlsx", "")
            dates.append(d)
        dates.sort(reverse=True)
        return dates

from config import settings
default_archive_dir = os.path.join(str(settings.DATA_DIR), "news_archive")
news_archive_service = ExcelNewsArchive(storage_dir=default_archive_dir)
