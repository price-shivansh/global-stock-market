"""
News History Module
Saves fetched news items to an Excel file as a persistent history log.
"""
import os
from datetime import datetime
from typing import List, Set
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from models import NewsItem

# Column definitions: (header label, attribute on NewsItem or custom, column width)
COLUMNS = [
    ("News Heading",     "title",           60),
    ("News Link",        "url",             50),
    ("Sentiment Effect", "sentiment",       18),
    ("Sentiment Score",  "sentiment_score", 16),
    ("Source",           "source",          20),
    ("Published Date",   "published",       22),
    ("Saved At",         "saved_at",        22),
]

HEADER_FILL   = PatternFill("solid", fgColor="1E293B")   # dark slate
HEADER_FONT   = Font(bold=True, color="FFFFFF", name="Calibri")
BULLISH_FILL  = PatternFill("solid", fgColor="D1FAE5")   # light green
BEARISH_FILL  = PatternFill("solid", fgColor="FEE2E2")   # light red
NEUTRAL_FILL  = PatternFill("solid", fgColor="F1F5F9")   # light gray


from config import settings

def get_history_path(category: str = "General") -> str:
    """Return the absolute path to the Excel file for a specific market category."""
    history_dir = str(settings.HISTORY_DIR)
    os.makedirs(history_dir, exist_ok=True)
    filename = f"{category.lower().replace(' ', '_')}_news_history.xlsx"
    return os.path.join(history_dir, filename)


def _load_or_create_workbook(category: str):
    """Return (workbook, sheet). Creates headers if the file is brand new."""
    excel_file = get_history_path(category)
    if os.path.exists(excel_file):
        wb = load_workbook(excel_file)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = f"{category} News"
        _write_headers(ws)
    return wb, ws


def _write_headers(ws):
    """Write styled column headers in row 1."""
    for col_idx, (label, _, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"         # freeze header row


def _get_existing_titles(ws) -> Set[str]:
    """Return a set of news titles already stored (skip header row)."""
    titles = set()
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0]:
            titles.add(str(row[0]).strip())
    return titles


def _row_fill(sentiment: str) -> PatternFill:
    s = str(sentiment).upper()
    if s == "BULLISH":
        return BULLISH_FILL
    if s == "BEARISH":
        return BEARISH_FILL
    return NEUTRAL_FILL


def save_news_to_excel(news_items: List[NewsItem]) -> int:
    """
    Append new (de-duplicated) news items to the Excel history file dynamically per category.
    Returns the total number of rows that were actually added across all categories.
    """
    if not news_items:
        return 0

    from collections import defaultdict
    category_items = defaultdict(list)
    for item in news_items:
        cat = getattr(item, "category", "General")
        category_items[cat].append(item)

    saved_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    total_added = 0

    for cat, items in category_items.items():
        wb, ws = _load_or_create_workbook(cat)
        existing_titles = _get_existing_titles(ws)
        added = 0

        for item in items:
            title = item.title.strip()
            if title in existing_titles:
                continue   # skip duplicate

            row_data = [
                title,
                item.url,
                str(item.sentiment).upper(),
                round(item.sentiment_score, 4),
                item.source,
                item.published.strftime("%Y-%m-%d %H:%M:%S"),
                saved_at,
            ]

            row_idx = ws.max_row + 1
            fill = _row_fill(item.sentiment)

            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = fill
                cell.alignment = Alignment(wrap_text=False, vertical="center")
                # Make the URL column a clickable hyperlink
                if col_idx == 2 and item.url:
                    cell.hyperlink = item.url
                    cell.font = Font(color="2563EB", underline="single", name="Calibri")
                else:
                    cell.font = Font(name="Calibri")

            existing_titles.add(title)
            added += 1

        if added > 0:
            excel_file = get_history_path(cat)
            wb.save(excel_file)
            print(f"[news_history] Saved {added} new news items to {excel_file}")
            
        total_added += added

    if total_added == 0:
        print("[news_history] No new news items to save (all duplicates).")

    return total_added


def get_history_stats(category: str = "General") -> dict:
    """Return simple stats about the stored history for given category."""
    excel_file = get_history_path(category)
    if not os.path.exists(excel_file):
        return {"total_rows": 0, "file_exists": False, "file_path": excel_file}

    wb = load_workbook(excel_file, read_only=True)
    ws = wb.active
    total = ws.max_row - 1  # subtract header row
    wb.close()

    return {
        "total_rows": max(0, total),
        "file_exists": True,
        "file_path": excel_file
    }

def get_all_history_stats() -> dict:
    categories = ["Indian Markets", "Crypto", "Commodities", "Global Markets", "General"]
    stats = {}
    for cat in categories:
        stats[cat] = get_history_stats(cat)
    return stats

def reset_history(category: str) -> bool:
    """Delete the history file for the specified category. Returns True if deleted or did not exist."""
    excel_file = get_history_path(category)
    if os.path.exists(excel_file):
        try:
            os.remove(excel_file)
            return True
        except Exception as e:
            print(f"Error deleting {excel_file}: {e}")
            return False
    return True
